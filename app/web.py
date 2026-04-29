from pathlib import Path
from typing import Optional
import json
import io
from datetime import datetime, timezone as _tz, date as _date
from collections import defaultdict

import base64
from fastapi import APIRouter, Request, Form, Depends, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy.orm import Session
from sqlalchemy import func

from .models import SessionLocal, Store, Member, Transaction, Receipt, BillingProfile, Product, StaffUser, StoreSubscription, SubscriptionPlan, SubscriptionInvoice, Company
from .core.security import hash_password, verify_password, create_access_token, decode_token, validate_password
from .services.member_service import MemberService

BASE_DIR = Path(__file__).parent
templates = Jinja2Templates(directory=str(BASE_DIR / "templates"))

# ── Number format filters (comma thousand separator) ──
def _fmt(v, d=2):
    try: return f"{float(v):,.{d}f}"
    except (TypeError, ValueError): return "0." + "0"*d
templates.env.filters["fmt"]  = lambda v: _fmt(v, 2)
templates.env.filters["fmt0"] = lambda v: _fmt(v, 0)
templates.env.filters["fmt1"] = lambda v: _fmt(v, 1)

router = APIRouter()


# ──────────────────────────────────────────
# Thai baht number-to-words helper
# ──────────────────────────────────────────
def thai_baht_text(amount: float) -> str:
    _ones = ['', 'หนึ่ง', 'สอง', 'สาม', 'สี่', 'ห้า', 'หก', 'เจ็ด', 'แปด', 'เก้า']

    def _digits(n: int) -> str:
        if n == 0:
            return 'ศูนย์'
        if n >= 1_000_000:
            return _digits(n // 1_000_000) + 'ล้าน' + (_digits(n % 1_000_000) if n % 1_000_000 else '')
        result = ''
        for level in range(5, -1, -1):
            digit = (n // (10 ** level)) % 10
            if digit == 0:
                continue
            if level == 1:
                if digit == 1:
                    result += 'สิบ'
                elif digit == 2:
                    result += 'ยี่สิบ'
                else:
                    result += _ones[digit] + 'สิบ'
            elif level == 0 and n >= 10 and digit == 1:
                result += 'เอ็ด'
            else:
                units = ['', 'สิบ', 'ร้อย', 'พัน', 'หมื่น', 'แสน']
                result += _ones[digit] + units[level]
        return result

    amount = round(float(amount), 2)
    baht_int = int(amount)
    satang = round((amount - baht_int) * 100)
    result = _digits(baht_int) + 'บาท'
    if satang:
        result += _digits(satang) + 'สตางค์'
    else:
        result += 'ถ้วน'
    return result


# ──────────────────────────────────────────
# DB helper
# ──────────────────────────────────────────
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# ──────────────────────────────────────────
# Auth helper: ดึง Store จาก cookie JWT
# ──────────────────────────────────────────
def _get_store(request: Request, db: Session) -> Optional[Store]:
    token = request.cookies.get("access_token")
    if not token:
        return None
    try:
        payload = decode_token(token)
        store_id = int(payload.get("sub", 0))
        return db.query(Store).filter(Store.id == store_id).first()
    except Exception:
        return None


def _auto_check_subscription(store: Store, db: Session):
    """Automatically transition expired subscriptions to grace/expired status."""
    if store.subscription_status not in ("active", "grace"):
        return
    sub = (
        db.query(StoreSubscription)
        .filter(
            StoreSubscription.store_id == store.id,
            StoreSubscription.status.in_(["active", "grace"]),
        )
        .order_by(StoreSubscription.id.desc())
        .first()
    )
    if not sub:
        return
    now = datetime.now(_tz.utc)
    changed = False
    if sub.status == "active" and sub.expires_at and now > sub.expires_at:
        sub.status = "grace"
        store.subscription_status = "grace"
        changed = True
    if sub.status == "grace" and sub.grace_until and now > sub.grace_until:
        sub.status = "expired"
        store.subscription_status = "expired"
        changed = True
    if changed:
        db.add(sub)
        db.add(store)
        db.commit()
        db.refresh(store)


def _get_user_ctx(request: Request, db: Session) -> Optional[dict]:
    """Return {'store', 'role', 'actor_id', 'actor_name'} or None if unauthenticated."""
    token = request.cookies.get("access_token")
    if not token:
        return None
    try:
        payload = decode_token(token)
        store_id = int(payload.get("sub", 0))
        store = db.query(Store).filter(Store.id == store_id).first()
        if not store:
            return None
        # Block suspended / rejected stores
        if getattr(store, "store_status", "active") in ("rejected", "suspended"):
            return None
        # Auto-check subscription expiry
        _auto_check_subscription(store, db)
        user_type = payload.get("user_type", "store")
        if user_type == "store":
            role = "owner"
            actor_id = store.id
            actor_name = f"{store.name} (เจ้าของ)"
        else:
            role = payload.get("staff_role", "user")
            actor_id = payload.get("staff_id")
            actor_name = payload.get("staff_name", "staff")
        return {"store": store, "role": role, "actor_id": actor_id, "actor_name": actor_name}
    except Exception:
        return None


def _is_admin_or_owner(ctx: dict) -> bool:
    return ctx["role"] in ("owner", "admin")


def _get_products(db: Session, store_id: int):
    return db.query(Product).filter(Product.store_id == store_id, Product.is_active == True).all()


# ──────────────────────────────────────────
# Usage limits helpers
# ──────────────────────────────────────────
_FREE_TIER = {
    "max_members": 50,
    "max_staff": 2,
    "max_receipts_per_month": 100,
    "max_products": 10,
}

STAFF_LIMIT = 10  # absolute max regardless of plan


def _get_plan_limits(store: Store, db: Session) -> dict:
    """Return limit dict from active subscription plan, or free-tier defaults."""
    if store.subscription_status not in ("active", "grace"):
        return _FREE_TIER.copy()
    sub = (
        db.query(StoreSubscription)
        .filter(
            StoreSubscription.store_id == store.id,
            StoreSubscription.status.in_(["active", "grace"]),
        )
        .order_by(StoreSubscription.id.desc())
        .first()
    )
    if not sub or not sub.plan_id:
        return _FREE_TIER.copy()
    plan = db.query(SubscriptionPlan).filter(SubscriptionPlan.id == sub.plan_id).first()
    if not plan:
        return _FREE_TIER.copy()
    return {
        "max_members": plan.max_members,
        "max_staff": plan.max_staff,
        "max_receipts_per_month": plan.max_receipts_per_month,
        "max_products": plan.max_products,
    }


def _check_limit(limits: dict, resource: str, current_count: int) -> tuple[bool, str]:
    """Return (allowed, error_message). 0 in limits means unlimited."""
    limit = limits.get(resource, 0)
    if limit == 0:
        return True, ""
    if current_count >= limit:
        label_map = {
            "max_members": "สมาชิก",
            "max_staff": "ผู้ใช้งาน",
            "max_receipts_per_month": "ใบเสร็จต่อเดือน",
            "max_products": "สินค้า",
        }
        label = label_map.get(resource, resource)
        return False, f"ถึงขีดจำกัด {label} ({current_count}/{limit}) ตามแผนปัจจุบัน กรุณาอัพเกรดแผน"
    return True, ""


# ──────────────────────────────────────────
# Login / Logout
# ──────────────────────────────────────────
@router.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    return templates.TemplateResponse(request, "login.html")


@router.post("/login")
def login_post(
    request: Request,
    username: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db),
):
    store = db.query(Store).filter(Store.username == username).first()
    if store and verify_password(password, store.hashed_password):
        # Block rejected/suspended stores
        if getattr(store, "store_status", "active") in ("rejected", "suspended"):
            reason = getattr(store, "rejection_reason", "") or "ไม่ระบุเหตุผล"
            label = "ถูกปฏิเสธ" if store.store_status == "rejected" else "ถูกระงับ"
            return templates.TemplateResponse(
                request, "login.html",
                {"error": f"บัญชีนี้{label}: {reason}"},
            )
        token = create_access_token({
            "sub": str(store.id),
            "store_name": store.name,
            "user_type": "store",
        })
    else:
        # ลองค้นหาจาก StaffUser
        staff = db.query(StaffUser).filter(
            StaffUser.username == username,
            StaffUser.is_active == True,
        ).first()
        if not staff or not verify_password(password, staff.hashed_password):
            return templates.TemplateResponse(
                request,
                "login.html",
                {"error": "ชื่อผู้ใช้หรือรหัสผ่านไม่ถูกต้อง"},
            )
        # Check if parent store is rejected/suspended
        parent_store = db.query(Store).filter(Store.id == staff.store_id).first()
        if parent_store and getattr(parent_store, "store_status", "active") in ("rejected", "suspended"):
            reason = getattr(parent_store, "rejection_reason", "") or "ไม่ระบุเหตุผล"
            return templates.TemplateResponse(
                request, "login.html",
                {"error": f"บัญชีร้านค้านี้ถูกปฏิเสธ: {reason}"},
            )
        token = create_access_token({
            "sub": str(staff.store_id),
            "store_name": staff.store.name,
            "user_type": "staff",
            "staff_id": staff.id,
            "staff_role": staff.role,
            "staff_name": staff.name,
        })
    resp = RedirectResponse("/web/dashboard", status_code=302)
    resp.set_cookie("access_token", token, httponly=True, samesite="lax", max_age=28800)
    return resp


@router.get("/logout")
def logout():
    resp = RedirectResponse("/web/login", status_code=302)
    resp.delete_cookie("access_token")
    return resp


# ──────────────────────────────────────────
# ลงทะเบียนร้านค้าใหม่
# ──────────────────────────────────────────
@router.get("/register", response_class=HTMLResponse)
def register_page(request: Request, db: Session = Depends(get_db)):
    plans = db.query(SubscriptionPlan).filter(SubscriptionPlan.is_active == True).order_by(SubscriptionPlan.id).all()
    return templates.TemplateResponse(request, "register.html", {"plans": plans})


@router.post("/register")
def register_post(
    request: Request,
    store_name: str = Form(...),
    username: str = Form(...),
    password: str = Form(...),
    password_confirm: str = Form(...),
    phone: str = Form(""),
    email: str = Form(""),
    address: str = Form(""),
    tax_id: str = Form(""),
    business_type: str = Form(""),
    requested_plan_id: str = Form(""),
    db: Session = Depends(get_db),
):
    plans = db.query(SubscriptionPlan).filter(SubscriptionPlan.is_active == True).order_by(SubscriptionPlan.id).all()

    def _err(msg: str):
        return templates.TemplateResponse(request, "register.html", {"error": msg, "plans": plans})

    if password != password_confirm:
        return _err("รหัสผ่านไม่ตรงกัน")
    pwd_error = validate_password(password)
    if pwd_error:
        return _err(pwd_error)
    if db.query(Store).filter(Store.username == username).first():
        return _err("ชื่อผู้ใช้นี้มีอยู่แล้ว")
    if db.query(StaffUser).filter(StaffUser.username == username).first():
        return _err("ชื่อผู้ใช้นี้มีอยู่แล้ว")

    plan_id: int | None = None
    if requested_plan_id:
        try:
            plan_id = int(requested_plan_id)
        except ValueError:
            plan_id = None

    store = Store(
        name=store_name.strip(),
        username=username.strip(),
        hashed_password=hash_password(password),
        phone=phone.strip() or None,
        email=email.strip() or None,
        address=address.strip() or None,
        tax_id=tax_id.strip() or None,
        business_type=business_type or None,
        requested_plan_id=plan_id,
        store_status="active",
        subscription_status="free",
    )
    db.add(store)
    db.commit()
    return RedirectResponse("/web/login?registered=1", status_code=302)


# ──────────────────────────────────────────
# รายชื่อสมาชิก
# ──────────────────────────────────────────
@router.get("/members", response_class=HTMLResponse)
def members_page(
    request: Request,
    q: str = None,
    enrolled: str = None,
    page: int = 1,
    db: Session = Depends(get_db),
):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    store = ctx["store"]
    role = ctx["role"]

    PAGE_SIZE = 20
    if page < 1:
        page = 1

    mtype = request.query_params.get("mtype", "person")  # "person" | "company" | "all"

    query = db.query(Member).filter(Member.store_id == store.id)
    if mtype in ("person", "company"):
        query = query.filter(
            (Member.member_type == mtype) | (Member.member_type == None)
            if mtype == "person"
            else Member.member_type == mtype
        )
    if q:
        query = query.filter(
            Member.name.ilike(f"%{q}%")
            | Member.phone.ilike(f"%{q}%")
            | Member.member_code.ilike(f"%{q}%")
            | Member.company_name.ilike(f"%{q}%")
            | Member.license_plate.ilike(f"%{q}%")
            | Member.driver_name.ilike(f"%{q}%")
        )
    total = query.count()
    total_pages = max(1, (total + PAGE_SIZE - 1) // PAGE_SIZE)
    if page > total_pages:
        page = total_pages
    members = query.order_by(Member.id.desc()).offset((page - 1) * PAGE_SIZE).limit(PAGE_SIZE).all()
    return templates.TemplateResponse(
        request,
        "members.html",
        {
            "store_name": store.name,
            "members": members,
            "total": total,
            "q": q or "",
            "enrolled": enrolled,
            "user_role": role,
            "page": page,
            "total_pages": total_pages,
            "mtype": mtype,
        },
    )


# ──────────────────────────────────────────
# สมัครสมาชิก
# ──────────────────────────────────────────
@router.get("/enroll", response_class=HTMLResponse)
def enroll_page(request: Request, db: Session = Depends(get_db)):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    store = ctx["store"]
    return templates.TemplateResponse(
        request, "enroll.html", {"store_name": store.name, "user_role": ctx["role"]}
    )


@router.post("/enroll")
def enroll_post(
    request: Request,
    member_type: str = Form("person"),
    name: str = Form(...),
    phone: str = Form(None),
    email: str = Form(None),
    member_code: str = Form(None),
    birthdate: str = Form(None),
    tier: str = Form("general"),
    company_name: str = Form(None),
    tax_id: str = Form(None),
    address: str = Form(None),
    db: Session = Depends(get_db),
):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    store = ctx["store"]
    role = ctx["role"]

    # Usage limit check
    limits = _get_plan_limits(store, db)
    member_count = db.query(Member).filter(Member.store_id == store.id).count()
    allowed, limit_msg = _check_limit(limits, "max_members", member_count)
    if not allowed:
        return templates.TemplateResponse(
            request, "enroll.html",
            {"store_name": store.name, "user_role": role, "error": limit_msg},
        )

    # ตรวจสอบเบอร์ซ้ำในร้านเดียวกัน
    if phone:
        existing = (
            db.query(Member)
            .filter(Member.phone == phone, Member.store_id == store.id)
            .first()
        )
        if existing:
            return templates.TemplateResponse(
                request,
                "enroll.html",
                {
                    "store_name": store.name,
                    "user_role": role,
                    "error": f"เบอร์ {phone} มีอยู่ในระบบแล้ว",
                    "form": {
                        "name": name,
                        "phone": phone,
                        "email": email,
                        "birthdate": birthdate,
                        "tier": tier,
                        "company_name": company_name,
                        "address": address,
                    },
                },
            )

    member = Member(
        member_type=member_type if member_type in ("person", "company") else "person",
        name=name,
        phone=phone or None,
        email=email or None,
        birthdate=birthdate or None,
        tier=tier,
        store_id=store.id,
        company_name=company_name.strip() if company_name and company_name.strip() else None,
        tax_id=tax_id.strip() if tax_id and tax_id.strip() else None,
        address=address.strip() if address and address.strip() else None,
    )
    db.add(member)
    db.flush()  # รับ id ก่อน commit

    # สร้างรหัสสมาชิกอัตโนมัติถ้าไม่ได้กรอกมา
    member.member_code = member_code.strip() if member_code and member_code.strip() else f"M{store.id:04d}{member.id:07d}"
    db.commit()

    return RedirectResponse(f"/web/members/{member.id}/new-bill", status_code=302)


# ──────────────────────────────────────────
# รายละเอียดสมาชิก
# ──────────────────────────────────────────
@router.get("/members/{member_id}", response_class=HTMLResponse)
def member_detail(request: Request, member_id: int, db: Session = Depends(get_db)):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    store = ctx["store"]
    role = ctx["role"]
    member = (
        db.query(Member)
        .filter(Member.id == member_id, Member.store_id == store.id)
        .first()
    )
    if not member:
        return RedirectResponse("/web/members", status_code=302)
    receipts = (
        db.query(Receipt)
        .join(Transaction)
        .filter(Transaction.member_id == member_id, Receipt.deleted_at == None)
        .order_by(Receipt.printed_at.desc())
        .limit(20)
        .all()
    )
    billing_profiles = db.query(BillingProfile).filter(BillingProfile.member_id == member_id).all()
    return templates.TemplateResponse(
        request,
        "member_detail.html",
        {"store_name": store.name, "member": member, "receipts": receipts, "billing_profiles": billing_profiles, "user_role": role},
    )


# ──────────────────────────────────────────
# ออกบิล / ใบเสร็จ
# ──────────────────────────────────────────
@router.get("/members/{member_id}/new-bill", response_class=HTMLResponse)
def new_bill_page(request: Request, member_id: int, db: Session = Depends(get_db)):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    store = ctx["store"]
    role = ctx["role"]
    member = (
        db.query(Member)
        .filter(Member.id == member_id, Member.store_id == store.id)
        .first()
    )
    if not member:
        return RedirectResponse("/web/members", status_code=302)
    billing_profiles = db.query(BillingProfile).filter(BillingProfile.member_id == member_id).all()
    products = _get_products(db, store.id)
    return templates.TemplateResponse(
        request, "new_bill.html", {
            "store_name": store.name,
            "store": store,
            "member": member,
            "billing_profiles": billing_profiles,
            "products": products,
            "user_role": role,
        }
    )


@router.post("/members/{member_id}/new-bill")
def new_bill_post(
    request: Request,
    member_id: int,
    items_json: str = Form(...),
    payment_method: str = Form("cash"),
    note: str = Form(""),
    billing_profile_id: str = Form(""),
    vat_type: str = Form("none"),
    new_bp_company: str = Form(""),
    new_bp_tax_id: str = Form(""),
    new_bp_address: str = Form(""),
    new_bp_phone: str = Form(""),
    new_bp_email: str = Form(""),
    new_bp_save: str = Form(""),
    driver_name: str = Form(""),
    license_plate: str = Form(""),
    db: Session = Depends(get_db),
):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    store = ctx["store"]
    member = (
        db.query(Member)
        .filter(Member.id == member_id, Member.store_id == store.id)
        .first()
    )
    if not member:
        return RedirectResponse("/web/members", status_code=302)

    # Usage limit: receipts per month
    limits = _get_plan_limits(store, db)
    now_utc = datetime.now(_tz.utc)
    month_start = now_utc.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    receipts_this_month = (
        db.query(Receipt)
        .join(Transaction)
        .filter(
            Transaction.member_id.in_(
                db.query(Member.id).filter(Member.store_id == store.id).scalar_subquery()
            ),
            Receipt.printed_at >= month_start,
            Receipt.deleted_at == None,
        )
        .count()
    )
    allowed, limit_msg = _check_limit(limits, "max_receipts_per_month", receipts_this_month)
    if not allowed:
        return RedirectResponse(f"/web/members/{member_id}?error=receipt_limit", status_code=302)

    items = json.loads(items_json)
    item_total = sum(float(i["qty"]) * float(i["price"]) for i in items)
    vat_rate = store.vat_rate or 7.0

    # VAT: none / exclusive (บวกเพิ่ม) / inclusive (รวมอยู่แล้ว)
    if vat_type == "exclusive":
        subtotal = round(item_total, 2)
        vat_amount = round(subtotal * vat_rate / 100, 2)
        total = round(subtotal + vat_amount, 2)
    elif vat_type == "inclusive":
        total = round(item_total, 2)
        subtotal = round(total / (1 + vat_rate / 100), 2)
        vat_amount = round(total - subtotal, 2)
    else:  # none
        subtotal = round(item_total, 2)
        vat_amount = 0.0
        vat_rate = 0.0
        total = subtotal

    points_earned = MemberService.points_for_amount(total)

    # Billing profile
    bp = None
    bp_snapshot = None
    if billing_profile_id == "__new__" and new_bp_company.strip():
        company = new_bp_company.strip()
        if new_bp_save:
            new_bp = BillingProfile(
                member_id=member_id,
                label=company,
                company_name=company,
                tax_id=new_bp_tax_id.strip() or None,
                address=new_bp_address.strip() or None,
                phone=new_bp_phone.strip() or None,
                is_default=False,
            )
            db.add(new_bp)
            db.flush()
        bp_snapshot = {
            "company_name": company,
            "tax_id": new_bp_tax_id.strip() or None,
            "address": new_bp_address.strip() or None,
            "phone": new_bp_phone.strip() or None,
            "email": new_bp_email.strip() or None,
            "label": company,
        }
    elif billing_profile_id:
        try:
            bp = db.query(BillingProfile).filter(
                BillingProfile.id == int(billing_profile_id),
                BillingProfile.member_id == member_id,
            ).first()
        except (ValueError, Exception):
            pass

    tx = Transaction(
        store_id=store.id,
        total=total,
        payment_method=payment_method,
        member_id=member_id,
        terminal_id="web",
        raw={"items": items, "subtotal": subtotal, "vat_amount": vat_amount,
             "vat_rate": vat_rate, "vat_type": vat_type, "total": total,
             "payment_method": payment_method, "note": note},
    )
    db.add(tx)
    db.commit()
    db.refresh(tx)

    bp_payload = bp_snapshot or (
        {
            "company_name": bp.company_name if bp else None,
            "tax_id": bp.tax_id if bp else None,
            "address": bp.address if bp else None,
            "phone": bp.phone if bp else None,
            "label": bp.label if bp else None,
        } if bp else None
    )

    receipt = Receipt(
        transaction_id=tx.id,
        raw_payload={
            "items": items,
            "subtotal": subtotal,
            "vat_amount": vat_amount,
            "vat_rate": vat_rate,
            "vat_type": vat_type,
            "total": total,
            "payment_method": payment_method,
            "note": note,
            "points_earned": points_earned,
            "billing_profile": bp_payload,
            "driver_name": driver_name.strip() if driver_name and driver_name.strip() else None,
            "license_plate": license_plate.strip().upper() if license_plate and license_plate.strip() else None,
        },
        created_by_name=ctx["actor_name"],
        created_by_id=ctx["actor_id"],
    )
    db.add(receipt)
    db.commit()
    db.refresh(receipt)

    MemberService.add_points(db, member_id, points_earned)

    return RedirectResponse(f"/web/receipts/{receipt.id}", status_code=302)


@router.get("/receipts/{receipt_id}", response_class=HTMLResponse)
def receipt_view(request: Request, receipt_id: int, db: Session = Depends(get_db)):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    store = ctx["store"]
    role = ctx["role"]
    receipt = db.query(Receipt).filter(Receipt.id == receipt_id).first()
    if not receipt:
        return RedirectResponse("/web/members", status_code=302)
    tx = receipt.transaction
    member = None
    if tx and tx.member_id:
        member = (
            db.query(Member)
            .filter(Member.id == tx.member_id, Member.store_id == store.id)
            .first()
        )
        if not member:
            return RedirectResponse("/web/members", status_code=302)
    payload = receipt.raw_payload or {}
    points_earned = payload.get("points_earned", 0)
    amount_text = thai_baht_text(payload.get("total", 0))
    is_cash = bool(payload.get("is_cash")) and payload.get("bill_type") != "vat"
    template_name = "receipt_cash.html" if is_cash else "receipt_view.html"
    return templates.TemplateResponse(
        request,
        template_name,
        {
            "store_name": store.name,
            "store": store,
            "receipt": receipt,
            "transaction": tx,
            "member": member,
            "payload": payload,
            "points_earned": points_earned,
            "amount_text": amount_text,
            "user_role": role,
            "is_deleted": receipt.deleted_at is not None,
        },
    )


# ──────────────────────────────────────────
# แก้ไข / ลบ ใบเสร็จ
# ──────────────────────────────────────────
@router.get("/receipts/{receipt_id}/edit", response_class=HTMLResponse)
def receipt_edit_page(request: Request, receipt_id: int, db: Session = Depends(get_db)):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    store = ctx["store"]
    role = ctx["role"]
    receipt = db.query(Receipt).filter(Receipt.id == receipt_id).first()
    if not receipt:
        return RedirectResponse("/web/members", status_code=302)
    # ไม่อนุญาตแก้บิลที่ถูกลบแล้ว
    if receipt.deleted_at:
        return RedirectResponse(f"/web/receipts/{receipt_id}", status_code=302)
    tx = receipt.transaction
    payload = receipt.raw_payload or {}
    is_vat_bill = payload.get("bill_type") == "vat"
    member = None
    if tx and tx.member_id:
        member = db.query(Member).filter(Member.id == tx.member_id, Member.store_id == store.id).first()
        if not member:
            return RedirectResponse("/web/members", status_code=302)
    elif not is_vat_bill:
        # Non-VAT, no member — shouldn't normally exist; redirect to safety
        return RedirectResponse("/web/members", status_code=302)
    billing_profiles = db.query(BillingProfile).filter(BillingProfile.member_id == tx.member_id).all() if (tx and tx.member_id) else []
    companies = db.query(Company).filter(Company.store_id == store.id).order_by(Company.company_name).all() if is_vat_bill else []
    bp_raw = payload.get("billing_profile")
    # Detect if the saved billing profile has been deleted (label no longer matches any active profile)
    bp_orphaned = False
    if bp_raw and bp_raw.get("label"):
        active_labels = {bp.label for bp in billing_profiles}
        if bp_raw["label"] not in active_labels:
            bp_orphaned = True
    products = _get_products(db, store.id)
    return templates.TemplateResponse(request, "receipt_edit.html", {
        "store_name": store.name,
        "store": store,
        "receipt": receipt,
        "member": member,
        "payload": payload,
        "billing_profiles": billing_profiles,
        "companies": companies,
        "is_vat_bill": is_vat_bill,
        "bp_orphaned": bp_orphaned,
        "products": products,
        "user_role": role,
    })


@router.post("/receipts/{receipt_id}/edit")
def receipt_edit_post(
    request: Request,
    receipt_id: int,
    items_json: str = Form(...),
    payment_method: str = Form("cash"),
    note: str = Form(""),
    billing_profile_id: str = Form(""),
    company_id: str = Form(""),
    vat_type: str = Form("none"),
    driver_name: str = Form(""),
    license_plate: str = Form(""),
    new_company_name: str = Form(""),
    new_company_tax_id: str = Form(""),
    new_company_address: str = Form(""),
    new_company_phone: str = Form(""),
    new_company_email: str = Form(""),
    new_company_save: str = Form(""),
    db: Session = Depends(get_db),
):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    store = ctx["store"]
    receipt = db.query(Receipt).filter(Receipt.id == receipt_id).first()
    if not receipt:
        return RedirectResponse("/web/members", status_code=302)
    if receipt.deleted_at:
        return RedirectResponse(f"/web/receipts/{receipt_id}", status_code=302)
    tx = receipt.transaction
    payload_check = receipt.raw_payload or {}
    is_vat_bill = payload_check.get("bill_type") == "vat"
    member = db.query(Member).filter(Member.id == tx.member_id, Member.store_id == store.id).first() if (tx and tx.member_id) else None
    if not member and not is_vat_bill:
        return RedirectResponse("/web/members", status_code=302)

    # บันทึก payload เก่าก่อนแก้ไข
    old_payload = dict(receipt.raw_payload or {})

    # Reverse old points (member bills only)
    old_points = (receipt.raw_payload or {}).get("points_earned", 0)
    if member:
        MemberService.add_points(db, member.id, -old_points)

    items = json.loads(items_json)
    item_total = sum(float(i["qty"]) * float(i["price"]) for i in items)
    vat_rate = store.vat_rate or 7.0
    if vat_type == "exclusive":
        subtotal = round(item_total, 2)
        vat_amount = round(subtotal * vat_rate / 100, 2)
        total = round(subtotal + vat_amount, 2)
    elif vat_type == "inclusive":
        total = round(item_total, 2)
        subtotal = round(total / (1 + vat_rate / 100), 2)
        vat_amount = round(total - subtotal, 2)
    else:
        subtotal = round(item_total, 2)
        vat_amount = 0.0
        vat_rate = 0.0
        total = subtotal

    points_earned = MemberService.points_for_amount(total) if member else 0

    bp = None
    kept_bp_snapshot = None
    if is_vat_bill:
        # VAT bill — resolve company from companies table
        if company_id == "__keep__":
            kept_bp_snapshot = (receipt.raw_payload or {}).get("billing_profile")
        elif company_id == "__new__" and new_company_name.strip():
            co_snapshot = {
                "company_name": new_company_name.strip(),
                "tax_id": new_company_tax_id.strip() or None,
                "address": new_company_address.strip() or None,
                "phone": new_company_phone.strip() or None,
                "email": new_company_email.strip() or None,
                "label": new_company_name.strip(),
            }
            kept_bp_snapshot = co_snapshot
            if new_company_save:
                co = Company(
                    store_id=store.id,
                    company_name=new_company_name.strip(),
                    tax_id=new_company_tax_id.strip() or None,
                    address=new_company_address.strip() or None,
                    phone=new_company_phone.strip() or None,
                    email=new_company_email.strip() or None,
                )
                db.add(co)
                db.flush()
        elif company_id:
            try:
                co = db.query(Company).filter(
                    Company.id == int(company_id),
                    Company.store_id == store.id,
                ).first()
                if co:
                    kept_bp_snapshot = {
                        "company_name": co.company_name,
                        "tax_id": co.tax_id,
                        "address": co.address,
                        "phone": co.phone,
                        "email": co.email,
                        "label": co.company_name,
                    }
            except (ValueError, Exception):
                pass
    else:
        # Member bill — existing billing_profile logic
        if billing_profile_id == "__keep__":
            kept_bp_snapshot = (receipt.raw_payload or {}).get("billing_profile")
        elif billing_profile_id:
            try:
                bp = db.query(BillingProfile).filter(
                    BillingProfile.id == int(billing_profile_id),
                    BillingProfile.member_id == member.id,
                ).first()
            except (ValueError, Exception):
                pass

    # Update transaction
    tx.total = total
    tx.payment_method = payment_method
    tx.raw = {"items": items, "subtotal": subtotal, "vat_amount": vat_amount,
              "vat_rate": vat_rate, "vat_type": vat_type, "total": total,
              "payment_method": payment_method, "note": note}
    # Update receipt
    new_driver_name = driver_name.strip() if driver_name and driver_name.strip() else None
    new_license_plate = license_plate.strip().upper() if license_plate and license_plate.strip() else None
    receipt.raw_payload = {
        "items": items, "subtotal": subtotal, "vat_amount": vat_amount,
        "vat_rate": vat_rate, "vat_type": vat_type, "total": total,
        "payment_method": payment_method, "note": note,
        "points_earned": points_earned,
        "driver_name": new_driver_name,
        "license_plate": new_license_plate,
        "bill_type": payload_check.get("bill_type"),
        "is_cash": payload_check.get("is_cash", False),
        "billing_profile": kept_bp_snapshot if kept_bp_snapshot else (
            {
                "company_name": bp.company_name if bp else None,
                "tax_id": bp.tax_id if bp else None,
                "address": bp.address if bp else None,
                "phone": bp.phone if bp else None,
                "email": getattr(bp, "email", None) if bp else None,
                "label": bp.label if bp else None,
            } if bp else None
        ),
    }

    # บันทึก edit log
    edit_changes: dict = {}
    if abs(old_payload.get("total", 0) - total) > 0.005:
        edit_changes["total"] = {"old": old_payload.get("total", 0), "new": total}
    if old_payload.get("payment_method", "") != payment_method:
        edit_changes["payment_method"] = {"old": old_payload.get("payment_method", ""), "new": payment_method}
    old_items = old_payload.get("items", [])
    if json.dumps(old_items, sort_keys=True) != json.dumps(items, sort_keys=True):
        edit_changes["items"] = {"old_count": len(old_items), "new_count": len(items)}
    if old_payload.get("driver_name", "") != (new_driver_name or ""):
        edit_changes["driver_name"] = {"old": old_payload.get("driver_name", ""), "new": new_driver_name or ""}
    if old_payload.get("license_plate", "") != (new_license_plate or ""):
        edit_changes["license_plate"] = {"old": old_payload.get("license_plate", ""), "new": new_license_plate or ""}
    edit_entry = {
        "at": datetime.now(_tz.utc).isoformat(),
        "by_name": ctx["actor_name"],
        "by_id": ctx["actor_id"],
        "changes": edit_changes,
    }
    current_log = list(receipt.edit_log or [])
    current_log.append(edit_entry)
    receipt.edit_log = current_log

    db.add(tx)
    db.add(receipt)
    db.commit()

    if member:
        MemberService.add_points(db, member.id, points_earned)

    return RedirectResponse(f"/web/receipts/{receipt_id}", status_code=302)


@router.post("/receipts/{receipt_id}/delete")
def receipt_delete(request: Request, receipt_id: int, db: Session = Depends(get_db)):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    if not _is_admin_or_owner(ctx):
        return RedirectResponse("/web/members", status_code=302)
    store = ctx["store"]
    receipt = db.query(Receipt).filter(Receipt.id == receipt_id).first()
    if not receipt:
        return RedirectResponse("/web/members", status_code=302)
    tx = receipt.transaction
    member_id = tx.member_id if tx else None
    # ตรวจสอบว่าเป็นสมาชิกของร้านนี้
    if member_id:
        member = db.query(Member).filter(Member.id == member_id, Member.store_id == store.id).first()
        if not member:
            return RedirectResponse("/web/members", status_code=302)
    # Reverse points
    old_points = (receipt.raw_payload or {}).get("points_earned", 0)
    if member_id and old_points:
        MemberService.add_points(db, member_id, -old_points)
    # Soft delete — เก็บประวัติการลบ
    receipt.deleted_at = datetime.now(_tz.utc)
    receipt.deleted_by_name = ctx["actor_name"]
    receipt.deleted_by_id = ctx["actor_id"]
    db.commit()
    if member_id:
        return RedirectResponse(f"/web/members/{member_id}", status_code=302)
    return RedirectResponse("/web/members", status_code=302)


# ──────────────────────────────────────────
# แก้ไขข้อมูลสมาชิก
# ──────────────────────────────────────────
@router.get("/members/{member_id}/edit", response_class=HTMLResponse)
def member_edit_page(request: Request, member_id: int, db: Session = Depends(get_db)):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    store = ctx["store"]
    role = ctx["role"]
    member = db.query(Member).filter(Member.id == member_id, Member.store_id == store.id).first()
    if not member:
        return RedirectResponse("/web/members", status_code=302)
    receipts_count = (
        db.query(Receipt)
        .join(Transaction)
        .filter(Transaction.member_id == member_id, Receipt.deleted_at == None)
        .count()
    )
    return templates.TemplateResponse(request, "edit_member.html", {
        "store_name": store.name, "member": member, "receipts_count": receipts_count, "user_role": role,
    })


@router.post("/members/{member_id}/edit")
def member_edit_post(
    request: Request,
    member_id: int,
    member_type: str = Form("person"),
    name: str = Form(...),
    phone: str = Form(None),
    email: str = Form(None),
    member_code: str = Form(None),
    birthdate: str = Form(None),
    tier: str = Form("general"),
    company_name: str = Form(None),
    tax_id: str = Form(None),
    address: str = Form(None),
    db: Session = Depends(get_db),
):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    store = ctx["store"]
    role = ctx["role"]
    member = db.query(Member).filter(Member.id == member_id, Member.store_id == store.id).first()
    if not member:
        return RedirectResponse("/web/members", status_code=302)
    # ตรวจสอบเบอร์ซ้ำ (ยกเว้นตัวเอง)
    if phone:
        dup = db.query(Member).filter(
            Member.phone == phone,
            Member.store_id == store.id,
            Member.id != member_id,
        ).first()
        if dup:
            return templates.TemplateResponse(request, "edit_member.html", {
                "store_name": store.name,
                "member": member,
                "user_role": role,
                "error": f"เบอร์ {phone} มีอยู่ในระบบแล้ว (สมาชิก {dup.name})",
            })
    member.member_type = member_type if member_type in ("person", "company") else "person"
    member.name = name
    member.phone = phone or None
    member.email = email or None
    member.member_code = member_code.strip() if member_code and member_code.strip() else member.member_code
    member.birthdate = birthdate or None
    member.tier = tier
    member.company_name = company_name.strip() if company_name and company_name.strip() else None
    member.tax_id = tax_id.strip() if tax_id and tax_id.strip() else None
    member.address = address.strip() if address and address.strip() else None
    db.commit()
    return RedirectResponse(f"/web/members/{member_id}?updated=1", status_code=302)


@router.post("/members/{member_id}/delete")
def member_delete(request: Request, member_id: int, db: Session = Depends(get_db)):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    if not _is_admin_or_owner(ctx):
        return RedirectResponse(f"/web/members/{member_id}", status_code=302)
    store = ctx["store"]
    member = db.query(Member).filter(Member.id == member_id, Member.store_id == store.id).first()
    if member:
        # ลบใบเสร็จและ transaction ที่เกี่ยวข้อง
        txs = db.query(Transaction).filter(Transaction.member_id == member_id).all()
        for tx in txs:
            db.query(Receipt).filter(Receipt.transaction_id == tx.id).delete()
        db.query(Transaction).filter(Transaction.member_id == member_id).delete()
        db.query(BillingProfile).filter(BillingProfile.member_id == member_id).delete()
        db.delete(member)
        db.commit()
    return RedirectResponse("/web/members?deleted=1", status_code=302)


# ──────────────────────────────────────────
# เพิ่มคะแนนด้วยมือ (staff action)
# ──────────────────────────────────────────
@router.post("/members/{member_id}/add-points")
def add_points_post(
    request: Request,
    member_id: int,
    points: int = Form(...),
    db: Session = Depends(get_db),
):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    store = ctx["store"]
    member = (
        db.query(Member)
        .filter(Member.id == member_id, Member.store_id == store.id)
        .first()
    )
    if member:
        member.points = (member.points or 0) + points
        db.commit()
    return RedirectResponse(f"/web/members/{member_id}?updated=1", status_code=302)


# ──────────────────────────────────────────
# ที่อยู่ออกบิล (Billing Profiles) ของสมาชิก
# ──────────────────────────────────────────
@router.get("/members/{member_id}/billing-profiles", response_class=HTMLResponse)
def billing_profiles_page(request: Request, member_id: int, db: Session = Depends(get_db)):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    store = ctx["store"]
    member = db.query(Member).filter(Member.id == member_id, Member.store_id == store.id).first()
    if not member:
        return RedirectResponse("/web/members", status_code=302)
    profiles = db.query(BillingProfile).filter(BillingProfile.member_id == member_id).all()
    return templates.TemplateResponse(request, "billing_profiles.html", {
        "store_name": store.name, "member": member, "profiles": profiles, "user_role": ctx["role"],
    })


@router.post("/members/{member_id}/update-company")
def update_company_info(
    request: Request, member_id: int,
    company_name: str = Form(""),
    tax_id: str = Form(""),
    address: str = Form(""),
    phone: str = Form(""),
    email: str = Form(""),
    set_default: str = Form(""),
    db: Session = Depends(get_db),
):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    store = ctx["store"]
    member = db.query(Member).filter(Member.id == member_id, Member.store_id == store.id, Member.member_type == "company").first()
    if not member:
        return RedirectResponse("/web/members", status_code=302)
    if company_name.strip():
        member.company_name = company_name.strip()
    member.tax_id = tax_id.strip() or None
    member.address = address.strip() or None
    member.phone = phone.strip() or None
    member.email = email.strip() or None
    if set_default:
        db.query(BillingProfile).filter(BillingProfile.member_id == member_id).update({"is_default": False})
    db.commit()
    return RedirectResponse(f"/web/members/{member_id}", status_code=302)


@router.post("/members/{member_id}/set-company-default")
def set_company_default(
    request: Request, member_id: int,
    db: Session = Depends(get_db),
):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    store = ctx["store"]
    member = db.query(Member).filter(Member.id == member_id, Member.store_id == store.id).first()
    if not member:
        return RedirectResponse("/web/members", status_code=302)
    # clear all billing profile defaults so member's own data becomes implicit default
    db.query(BillingProfile).filter(BillingProfile.member_id == member_id).update({"is_default": False})
    db.commit()
    return RedirectResponse(f"/web/members/{member_id}", status_code=302)


@router.post("/members/{member_id}/clear-company-info")
def clear_company_info(
    request: Request, member_id: int,
    db: Session = Depends(get_db),
):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    store = ctx["store"]
    member = db.query(Member).filter(Member.id == member_id, Member.store_id == store.id, Member.member_type == "company").first()
    if not member:
        return RedirectResponse("/web/members", status_code=302)
    member.tax_id = None
    member.address = None
    member.phone = None
    member.email = None
    db.commit()
    return RedirectResponse(f"/web/members/{member_id}", status_code=302)


@router.post("/members/{member_id}/billing-profiles")
def billing_profiles_post(
    request: Request, member_id: int,
    company_name: str = Form(...),
    label: str = Form(""),
    tax_id: str = Form(""),
    address: str = Form(""),
    phone: str = Form(""),
    email: str = Form(""),
    is_default: str = Form(""),
    db: Session = Depends(get_db),
):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    store = ctx["store"]
    member = db.query(Member).filter(Member.id == member_id, Member.store_id == store.id).first()
    if not member:
        return RedirectResponse("/web/members", status_code=302)
    effective_label = label.strip() or company_name.strip()
    if is_default:
        db.query(BillingProfile).filter(BillingProfile.member_id == member_id).update({"is_default": False})
    bp = BillingProfile(
        member_id=member_id, label=effective_label,
        company_name=company_name or None, tax_id=tax_id or None,
        address=address or None, phone=phone or None,
        email=email or None, is_default=bool(is_default),
    )
    db.add(bp)
    db.commit()
    return RedirectResponse(f"/web/members/{member_id}/billing-profiles?saved=1", status_code=302)


@router.post("/members/{member_id}/billing-profiles/{bp_id}/delete")
def billing_profile_delete(request: Request, member_id: int, bp_id: int, db: Session = Depends(get_db)):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    store = ctx["store"]
    bp = db.query(BillingProfile).filter(BillingProfile.id == bp_id, BillingProfile.member_id == member_id).first()
    if bp:
        db.delete(bp)
        db.commit()
    return RedirectResponse(f"/web/members/{member_id}/billing-profiles", status_code=302)


@router.post("/members/{member_id}/billing-profiles/{bp_id}/edit")
def billing_profile_edit(
    request: Request, member_id: int, bp_id: int,
    company_name: str = Form(...),
    label: str = Form(""),
    tax_id: str = Form(""),
    address: str = Form(""),
    phone: str = Form(""),
    email: str = Form(""),
    is_default: str = Form(""),
    db: Session = Depends(get_db),
):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    store = ctx["store"]
    bp = db.query(BillingProfile).filter(
        BillingProfile.id == bp_id,
        BillingProfile.member_id == member_id,
    ).first()
    if not bp:
        return RedirectResponse(f"/web/members/{member_id}/billing-profiles", status_code=302)
    effective_label = label.strip() or company_name.strip()
    if is_default:
        db.query(BillingProfile).filter(
            BillingProfile.member_id == member_id,
            BillingProfile.id != bp_id,
        ).update({"is_default": False})
    bp.label = effective_label
    bp.company_name = company_name or None
    bp.tax_id = tax_id or None
    bp.address = address or None
    bp.phone = phone or None
    bp.email = email or None
    bp.is_default = bool(is_default)
    db.commit()
    return RedirectResponse(f"/web/members/{member_id}/billing-profiles?saved=1", status_code=302)


# ──────────────────────────────────────────
# ตั้งค่าร้านค้า (Store Settings)
# ──────────────────────────────────────────
@router.get("/settings", response_class=HTMLResponse)
def settings_page(request: Request, db: Session = Depends(get_db)):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    if not _is_admin_or_owner(ctx):
        return RedirectResponse("/web/members", status_code=302)
    store = ctx["store"]
    return templates.TemplateResponse(request, "store_settings.html", {
        "store_name": store.name, "store": store, "user_role": ctx["role"],
    })


@router.post("/settings/profile")
def settings_profile_post(
    request: Request,
    store_name: str = Form(...),
    address: str = Form(""),
    tax_id: str = Form(""),
    phone: str = Form(""),
    email: str = Form(""),
    vat_rate: float = Form(7.0),
    include_vat: str = Form(""),
    receipt_color: str = Form(""),
    cash_receipt_color: str = Form(""),
    receipt_header_text: str = Form(""),
    receipt_footer_text: str = Form(""),
    db: Session = Depends(get_db),
):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    if not _is_admin_or_owner(ctx):
        return RedirectResponse("/web/members", status_code=302)
    store = ctx["store"]
    store.name = store_name
    store.address = address or None
    store.tax_id = tax_id or None
    store.phone = phone or None
    store.email = email or None
    store.vat_rate = vat_rate
    store.include_vat = bool(include_vat)
    store.receipt_color = receipt_color.strip() or None
    store.cash_receipt_color = cash_receipt_color.strip() or None
    store.receipt_header_text = receipt_header_text.strip() or None
    store.receipt_footer_text = receipt_footer_text.strip() or None
    db.commit()
    return RedirectResponse("/web/settings?saved=1", status_code=302)


@router.post("/settings/logo")
async def settings_logo_post(
    request: Request,
    logo: UploadFile = File(None),
    remove_logo: str = Form(""),
    db: Session = Depends(get_db),
):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    if not _is_admin_or_owner(ctx):
        return RedirectResponse("/web/members", status_code=302)
    store = ctx["store"]
    if remove_logo:
        store.logo_base64 = None
    elif logo and logo.filename:
        content = await logo.read()
        if len(content) > 500 * 1024:
            return RedirectResponse("/web/settings?logo_error=1", status_code=302)
        mime = logo.content_type or "image/png"
        if mime not in ("image/png", "image/jpeg", "image/gif", "image/webp", "image/svg+xml"):
            return RedirectResponse("/web/settings?logo_error=1", status_code=302)
        store.logo_base64 = f"data:{mime};base64,{base64.b64encode(content).decode()}"
    db.commit()
    return RedirectResponse("/web/settings?saved=1", status_code=302)


@router.post("/settings/staff/add")
def settings_staff_add(
    request: Request,
    name: str = Form(...),
    username: str = Form(...),
    password: str = Form(...),
    db: Session = Depends(get_db),
):
    # Route kept for backward compat — redirects to /web/staff
    return RedirectResponse("/web/staff", status_code=302)


@router.post("/settings/staff/{staff_id}/delete")
def settings_staff_delete(request: Request, staff_id: int, db: Session = Depends(get_db)):
    # Route kept for backward compat — redirects to /web/staff
    return RedirectResponse("/web/staff", status_code=302)


# ──────────────────────────────────────────
# รายการสินค้า (Products)
# ──────────────────────────────────────────
@router.get("/products", response_class=HTMLResponse)
def products_page(request: Request, db: Session = Depends(get_db)):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    store = ctx["store"]
    products = db.query(Product).filter(Product.store_id == store.id).order_by(Product.category, Product.name).all()
    return templates.TemplateResponse(request, "products.html", {
        "store_name": store.name, "products": products, "user_role": ctx["role"],
    })


@router.post("/products/batch-price")
def product_batch_price(
    request: Request,
    prices_json: str = Form(...),
    db: Session = Depends(get_db),
):
    """อัพเดทราคาหลายรายการพร้อมกัน — ทุก role ทำได้"""
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    store = ctx["store"]
    try:
        prices: dict = json.loads(prices_json)  # {str(product_id): price}
    except (ValueError, TypeError):
        return RedirectResponse("/web/products?saved=1", status_code=302)
    for pid_str, price_val in prices.items():
        try:
            pid = int(pid_str)
            price = float(price_val)
            if price < 0:
                continue
            p = db.query(Product).filter(Product.id == pid, Product.store_id == store.id).first()
            if p:
                p.price = round(price, 2)
        except (ValueError, TypeError):
            continue
    db.commit()
    return RedirectResponse("/web/products?saved=1", status_code=302)


@router.post("/products/add")
def product_add(
    request: Request,
    name: str = Form(...),
    unit: str = Form("ลิตร"),
    price: float = Form(0.0),
    category: str = Form("fuel"),
    db: Session = Depends(get_db),
):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    if not _is_admin_or_owner(ctx):
        return RedirectResponse("/web/products", status_code=302)
    store = ctx["store"]
    # Usage limit check
    limits = _get_plan_limits(store, db)
    product_count = db.query(Product).filter(Product.store_id == store.id, Product.is_active == True).count()
    allowed, limit_msg = _check_limit(limits, "max_products", product_count)
    if not allowed:
        return RedirectResponse(f"/web/products?limit_error=1", status_code=302)
    p = Product(store_id=store.id, name=name, unit=unit, price=price, category=category)
    db.add(p)
    db.commit()
    return RedirectResponse("/web/products?saved=1", status_code=302)


@router.post("/products/{product_id}/edit")
def product_edit(
    request: Request,
    product_id: int,
    name: str = Form(...),
    unit: str = Form("ลิตร"),
    price: float = Form(0.0),
    category: str = Form("fuel"),
    is_active: str = Form(""),
    db: Session = Depends(get_db),
):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    store = ctx["store"]
    p = db.query(Product).filter(Product.id == product_id, Product.store_id == store.id).first()
    if p:
        # user role can only update price; admin/owner can update all fields
        if _is_admin_or_owner(ctx):
            p.name = name; p.unit = unit; p.category = category; p.is_active = bool(is_active)
        p.price = price
        db.commit()
    return RedirectResponse("/web/products?saved=1", status_code=302)


@router.post("/products/{product_id}/delete")
def product_delete(request: Request, product_id: int, db: Session = Depends(get_db)):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    if not _is_admin_or_owner(ctx):
        return RedirectResponse("/web/products", status_code=302)
    store = ctx["store"]
    p = db.query(Product).filter(Product.id == product_id, Product.store_id == store.id).first()
    if p:
        db.delete(p)
        db.commit()
    return RedirectResponse("/web/products", status_code=302)


# ──────────────────────────────────────────
# จัดการผู้ใช้งาน (Staff Management)  — admin / owner only
# ──────────────────────────────────────────


@router.get("/staff", response_class=HTMLResponse)
def staff_list_page(request: Request, db: Session = Depends(get_db)):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    if not _is_admin_or_owner(ctx):
        return RedirectResponse("/web/members", status_code=302)
    store = ctx["store"]
    staff_list = db.query(StaffUser).filter(StaffUser.store_id == store.id).order_by(StaffUser.id).all()
    return templates.TemplateResponse(request, "staff.html", {
        "store_name": store.name,
        "user_role": ctx["role"],
        "staff_list": staff_list,
        "staff_count": len(staff_list),
        "staff_limit": STAFF_LIMIT,
    })


@router.get("/staff/new", response_class=HTMLResponse)
def staff_new_page(request: Request, db: Session = Depends(get_db)):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    if not _is_admin_or_owner(ctx):
        return RedirectResponse("/web/members", status_code=302)
    store = ctx["store"]
    staff_count = db.query(StaffUser).filter(StaffUser.store_id == store.id).count()
    return templates.TemplateResponse(request, "staff_form.html", {
        "store_name": store.name,
        "user_role": ctx["role"],
        "staff_count": staff_count,
        "staff_limit": STAFF_LIMIT,
        "action": "new",
    })


@router.post("/staff/generate")
def staff_generate(request: Request, db: Session = Depends(get_db)):
    """สร้าง staff users อัตโนมัติให้ครบ STAFF_LIMIT คน"""
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    if not _is_admin_or_owner(ctx):
        return RedirectResponse("/web/members", status_code=302)
    store = ctx["store"]

    existing = db.query(StaffUser).filter(StaffUser.store_id == store.id).all()
    current_count = len(existing)
    # Enforce plan limit
    limits = _get_plan_limits(store, db)
    plan_max = limits.get("max_staff", 0)
    effective_max = min(STAFF_LIMIT, plan_max) if plan_max > 0 else STAFF_LIMIT
    slots_left = effective_max - current_count
    if slots_left <= 0:
        return RedirectResponse("/web/staff?full=1", status_code=302)

    # หา prefix สั้น จาก store username (ไม่เกิน 6 ตัวอักษร)
    prefix = store.username[:6].lower()
    generated = []

    for n in range(1, STAFF_LIMIT + 1):
        if len(generated) >= slots_left:
            break
        password = f"{prefix}{n:02d}"
        # ตรวจว่า username ซ้ำไหม
        if db.query(StaffUser).filter(StaffUser.username == username).first():
            continue
        if db.query(Store).filter(Store.username == username).first():
            continue
        su = StaffUser(
            store_id=store.id,
            name=f"User {n:02d}",
            username=username,
            hashed_password=hash_password(password),
            role="user",
        )
        db.add(su)
        generated.append({"name": f"User {n:02d}", "username": username, "password": password})

    db.commit()

    # โหลด staff list ใหม่
    staff_list = db.query(StaffUser).filter(StaffUser.store_id == store.id).order_by(StaffUser.id).all()
    return templates.TemplateResponse(request, "staff.html", {
        "store_name": store.name,
        "user_role": ctx["role"],
        "staff_list": staff_list,
        "staff_count": len(staff_list),
        "staff_limit": STAFF_LIMIT,
        "generated_creds": generated,
    })


@router.post("/staff/new")
def staff_new_post(
    request: Request,
    name: str = Form(...),
    username: str = Form(...),
    password: str = Form(...),
    role: str = Form("user"),
    db: Session = Depends(get_db),
):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    if not _is_admin_or_owner(ctx):
        return RedirectResponse("/web/members", status_code=302)
    store = ctx["store"]

    if role not in ("admin", "user"):
        role = "user"

    staff_count = db.query(StaffUser).filter(StaffUser.store_id == store.id).count()
    limits = _get_plan_limits(store, db)
    plan_max = limits.get("max_staff", 0)
    effective_staff_limit = min(STAFF_LIMIT, plan_max) if plan_max > 0 else STAFF_LIMIT

    def _render_new_form(error: str):
        return templates.TemplateResponse(request, "staff_form.html", {
            "store_name": store.name,
            "user_role": ctx["role"],
            "staff_count": staff_count,
            "staff_limit": effective_staff_limit,
            "action": "new",
            "form": {"name": name, "username": username, "role": role},
            "error": error,
        })

    if staff_count >= effective_staff_limit:
        return _render_new_form(f"ถึงจำนวนผู้ใช้งานสูงสุดแล้ว ({effective_staff_limit} คน)")

    if db.query(StaffUser).filter(StaffUser.username == username).first() or \
       db.query(Store).filter(Store.username == username).first():
        return _render_new_form(f"username '{username}' ถูกใช้งานแล้ว")

    pwd_error = validate_password(password)
    if pwd_error:
        return _render_new_form(pwd_error)

    su = StaffUser(
        store_id=store.id,
        name=name,
        username=username,
        hashed_password=hash_password(password),
        role=role,
    )
    db.add(su)
    db.commit()
    return RedirectResponse("/web/staff?added=1", status_code=302)


@router.get("/staff/{staff_id}/edit", response_class=HTMLResponse)
def staff_edit_page(request: Request, staff_id: int, db: Session = Depends(get_db)):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    if not _is_admin_or_owner(ctx):
        return RedirectResponse("/web/members", status_code=302)
    store = ctx["store"]
    su = db.query(StaffUser).filter(StaffUser.id == staff_id, StaffUser.store_id == store.id).first()
    if not su:
        return RedirectResponse("/web/staff", status_code=302)
    return templates.TemplateResponse(request, "staff_form.html", {
        "store_name": store.name,
        "user_role": ctx["role"],
        "action": "edit",
        "staff": su,
        "form": {"name": su.name, "username": su.username, "role": su.role, "is_active": su.is_active},
    })


@router.post("/staff/{staff_id}/edit")
def staff_edit_post(
    request: Request,
    staff_id: int,
    name: str = Form(...),
    username: str = Form(...),
    password: str = Form(""),
    role: str = Form("user"),
    is_active: str = Form(""),
    db: Session = Depends(get_db),
):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    if not _is_admin_or_owner(ctx):
        return RedirectResponse("/web/members", status_code=302)
    store = ctx["store"]
    su = db.query(StaffUser).filter(StaffUser.id == staff_id, StaffUser.store_id == store.id).first()
    if not su:
        return RedirectResponse("/web/staff", status_code=302)

    if role not in ("admin", "user"):
        role = "user"

    def _render_edit_form(error: str):
        return templates.TemplateResponse(request, "staff_form.html", {
            "store_name": store.name,
            "user_role": ctx["role"],
            "action": "edit",
            "staff": su,
            "form": {"name": name, "username": username, "role": role, "is_active": bool(is_active)},
            "error": error,
        })

    dup_staff = db.query(StaffUser).filter(StaffUser.username == username, StaffUser.id != staff_id).first()
    dup_store = db.query(Store).filter(Store.username == username).first()
    if dup_staff or dup_store:
        return _render_edit_form(f"username '{username}' ถูกใช้งานแล้ว")

    if password and len(password) < 6:
        return _render_edit_form("รหัสผ่านต้องมีอย่างน้อย 6 ตัวอักษร")

    su.name = name
    su.username = username
    su.role = role
    su.is_active = bool(is_active)
    if password:
        su.hashed_password = hash_password(password)
    db.commit()
    return RedirectResponse("/web/staff?updated=1", status_code=302)


@router.post("/staff/{staff_id}/delete")
def staff_delete(request: Request, staff_id: int, db: Session = Depends(get_db)):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    if not _is_admin_or_owner(ctx):
        return RedirectResponse("/web/members", status_code=302)
    store = ctx["store"]
    su = db.query(StaffUser).filter(StaffUser.id == staff_id, StaffUser.store_id == store.id).first()
    if su:
        su.is_active = False  # soft-delete
        db.commit()
    return RedirectResponse("/web/staff?deleted=1", status_code=302)


# ──────────────────────────────────────────
# หน้าสรุปยอดขาย (Summary / Dashboard)
# ──────────────────────────────────────────

def _summary_query(db: Session, store_id: int, start: datetime, end: datetime):
    """คืนค่า list ของ (timestamp, total, payment_method) สำหรับใบเสร็จที่ไม่ถูกลบ"""
    return (
        db.query(Transaction.timestamp, Transaction.total, Transaction.payment_method)
        .join(Receipt, Receipt.transaction_id == Transaction.id)
        .filter(
            Transaction.store_id == store_id,
            Receipt.deleted_at == None,
            Transaction.timestamp >= start,
            Transaction.timestamp < end,
        )
        .all()
    )


# ──────────────────────────────────────────
# Subscription page (store-facing)
# ──────────────────────────────────────────
@router.get("/subscription", response_class=HTMLResponse)
def subscription_page(request: Request, db: Session = Depends(get_db)):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    store = ctx["store"]
    limits = _get_plan_limits(store, db)

    # Active subscription
    active_sub = (
        db.query(StoreSubscription)
        .filter(
            StoreSubscription.store_id == store.id,
            StoreSubscription.status.in_(["active", "grace"]),
        )
        .order_by(StoreSubscription.id.desc())
        .first()
    )
    active_plan = None
    if active_sub:
        active_plan = db.query(SubscriptionPlan).filter(SubscriptionPlan.id == active_sub.plan_id).first()

    # Invoice history
    invoices = (
        db.query(SubscriptionInvoice)
        .filter(SubscriptionInvoice.store_id == store.id)
        .order_by(SubscriptionInvoice.created_at.desc())
        .all()
    )

    # Current usage counts
    now_utc = datetime.now(_tz.utc)
    month_start = now_utc.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    member_count = db.query(Member).filter(Member.store_id == store.id).count()
    staff_count = db.query(StaffUser).filter(StaffUser.store_id == store.id).count()
    product_count = db.query(Product).filter(Product.store_id == store.id, Product.is_active == True).count()
    receipts_month = (
        db.query(Receipt)
        .join(Transaction)
        .filter(
            Transaction.member_id.in_(
                db.query(Member.id).filter(Member.store_id == store.id).scalar_subquery()
            ),
            Receipt.printed_at >= month_start,
            Receipt.deleted_at == None,
        )
        .count()
    )

    usage = {
        "members": {"current": member_count, "max": limits["max_members"]},
        "staff": {"current": staff_count, "max": limits["max_staff"]},
        "products": {"current": product_count, "max": limits["max_products"]},
        "receipts_month": {"current": receipts_month, "max": limits["max_receipts_per_month"]},
    }

    return templates.TemplateResponse(request, "subscription.html", {
        "store_name": store.name,
        "store": store,
        "user_role": ctx["role"],
        "active_sub": active_sub,
        "active_plan": active_plan,
        "invoices": invoices,
        "usage": usage,
    })


@router.get("/summary", response_class=HTMLResponse)
def summary_page(
    request: Request,
    period: str = "monthly",
    year: int = None,
    month: int = None,
    db: Session = Depends(get_db),
):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    store = ctx["store"]

    now = datetime.now()
    if not year:
        year = now.year
    if not month:
        month = now.month

    # กำหนดช่วงวันที่
    if period == "daily":
        start = datetime(year, month, 1)
        if month == 12:
            end = datetime(year + 1, 1, 1)
        else:
            end = datetime(year, month + 1, 1)
    elif period == "yearly":
        start = datetime(year - 4, 1, 1)
        end = datetime(year + 1, 1, 1)
    else:  # monthly (default)
        period = "monthly"
        start = datetime(year, 1, 1)
        end = datetime(year + 1, 1, 1)

    rows = _summary_query(db, store.id, start, end)

    # จัดกลุ่มด้วย Python
    grouped: dict = defaultdict(lambda: {"count": 0, "total": 0.0})
    for row in rows:
        ts = row.timestamp
        if not ts:
            continue
        if period == "daily":
            key = ts.strftime("%Y-%m-%d")
        elif period == "yearly":
            key = ts.strftime("%Y")
        else:
            key = ts.strftime("%Y-%m")
        grouped[key]["count"] += 1
        grouped[key]["total"] += row.total or 0.0

    sorted_data = sorted(grouped.items())
    total_bills = sum(v["count"] for _, v in sorted_data)
    total_amount = sum(v["total"] for _, v in sorted_data)

    # labels/values สำหรับ Chart.js
    chart_labels = [k for k, _ in sorted_data]
    chart_counts = [v["count"] for _, v in sorted_data]
    chart_totals = [round(v["total"], 2) for _, v in sorted_data]

    # ปีทั้งหมดที่มีข้อมูล (สำหรับ dropdown ย้อนหลัง)
    years_available = list(range(now.year, now.year - 5, -1))

    return templates.TemplateResponse(request, "summary.html", {
        "store_name": store.name,
        "user_role": ctx["role"],
        "period": period,
        "year": year,
        "month": month,
        "sorted_data": sorted_data,
        "total_bills": total_bills,
        "total_amount": round(total_amount, 2),
        "chart_labels": json.dumps(chart_labels, ensure_ascii=False),
        "chart_counts": json.dumps(chart_counts),
        "chart_totals": json.dumps(chart_totals),
        "years_available": years_available,
    })


@router.get("/summary/detail", response_class=HTMLResponse)
def summary_detail_page(
    request: Request,
    period: str = "monthly",
    date_key: str = "",
    db: Session = Depends(get_db),
):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    store = ctx["store"]

    # แปลง date_key เป็น start/end datetime
    try:
        if period == "daily" and len(date_key) == 10:          # YYYY-MM-DD
            start = datetime.strptime(date_key, "%Y-%m-%d")
            end = datetime(start.year, start.month, start.day + 1) if start.day < 28 else \
                  start.replace(day=1) + __import__("calendar").timedelta(days=32)
            # simpler: use timedelta
            from datetime import timedelta
            end = start + timedelta(days=1)
        elif period == "monthly" and len(date_key) == 7:       # YYYY-MM
            y, m = int(date_key[:4]), int(date_key[5:7])
            start = datetime(y, m, 1)
            end = datetime(y + 1, 1, 1) if m == 12 else datetime(y, m + 1, 1)
        elif period == "yearly" and len(date_key) == 4:        # YYYY
            y = int(date_key)
            start = datetime(y, 1, 1)
            end = datetime(y + 1, 1, 1)
        else:
            return RedirectResponse("/web/summary", status_code=302)
    except (ValueError, TypeError):
        return RedirectResponse("/web/summary", status_code=302)

    rows = (
        db.query(Receipt, Transaction)
        .join(Transaction, Transaction.id == Receipt.transaction_id)
        .filter(
            Transaction.store_id == store.id,
            Receipt.deleted_at == None,
            Transaction.timestamp >= start,
            Transaction.timestamp < end,
        )
        .order_by(Transaction.timestamp.desc())
        .all()
    )

    receipts_detail = []
    for receipt, tx in rows:
        payload = receipt.raw_payload or {}
        # หาสมาชิก (ถ้ามี)
        member = db.query(Member).filter(Member.id == tx.member_id).first() if tx.member_id else None
        bp = payload.get("billing_profile") or {}
        display_name = (
            bp.get("company_name")
            or (member.name if member else None)
            or payload.get("customer_name")
            or "–"
        )
        receipts_detail.append({
            "receipt": receipt,
            "transaction": tx,
            "member": member,
            "display_name": display_name,
            "total": tx.total,
            "payment_method": tx.payment_method or "–",
            "items_count": len(payload.get("items", [])),
            "created_by_name": receipt.created_by_name or "–",
        })

    total_amount = sum(r["total"] for r in receipts_detail)

    return templates.TemplateResponse(request, "summary_detail.html", {
        "store_name": store.name,
        "user_role": ctx["role"],
        "period": period,
        "date_key": date_key,
        "receipts_detail": receipts_detail,
        "total_amount": round(total_amount, 2),
        "total_bills": len(receipts_detail),
        "back_url": f"/web/summary?period={period}&year={date_key[:4]}",
    })


@router.get("/summary/export")
def summary_export(
    request: Request,
    period: str = "monthly",
    year: int = None,
    month: int = None,
    db: Session = Depends(get_db),
):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    store = ctx["store"]

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    now = datetime.now()
    if not year:
        year = now.year
    if not month:
        month = now.month

    if period == "daily":
        start = datetime(year, month, 1)
        end = datetime(year + 1, 1, 1) if month == 12 else datetime(year, month + 1, 1)
        period_label = f"รายวัน เดือน {month:02d}/{year}"
    elif period == "yearly":
        start = datetime(year - 4, 1, 1)
        end = datetime(year + 1, 1, 1)
        period_label = f"รายปี (5 ปีล่าสุด ถึงปี {year})"
    else:
        period = "monthly"
        start = datetime(year, 1, 1)
        end = datetime(year + 1, 1, 1)
        period_label = f"รายเดือน ปี {year}"

    rows = _summary_query(db, store.id, start, end)

    grouped: dict = defaultdict(lambda: {"count": 0, "total": 0.0})
    for row in rows:
        ts = row.timestamp
        if not ts:
            continue
        if period == "daily":
            key = ts.strftime("%Y-%m-%d")
        elif period == "yearly":
            key = ts.strftime("%Y")
        else:
            key = ts.strftime("%Y-%m")
        grouped[key]["count"] += 1
        grouped[key]["total"] += row.total or 0.0
    sorted_data = sorted(grouped.items())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "สรุปยอดขาย"

    # Title
    ws.merge_cells("A1:C1")
    title_cell = ws["A1"]
    title_cell.value = f"สรุปยอดขาย — {store.name}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:C2")
    sub_cell = ws["A2"]
    sub_cell.value = period_label
    sub_cell.alignment = Alignment(horizontal="center")

    ws.append([])

    # Header row
    header_fill = PatternFill("solid", fgColor="2563EB")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["ช่วงเวลา", "จำนวนใบเสร็จ", "ยอดรวม (บาท)"]
    ws.append(headers)
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    for key, v in sorted_data:
        ws.append([key, v["count"], round(v["total"], 2)])
        for col_idx in range(1, 4):
            ws.cell(row=ws.max_row, column=col_idx).border = border

    # Total row
    total_bills = sum(v["count"] for _, v in sorted_data)
    total_amount = round(sum(v["total"] for _, v in sorted_data), 2)
    ws.append(["รวมทั้งหมด", total_bills, total_amount])
    for col_idx in range(1, 4):
        cell = ws.cell(row=ws.max_row, column=col_idx)
        cell.font = Font(bold=True)
        cell.border = border

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"summary_{period}_{year}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ──────────────────────────────────────────
# Store owner dashboard
# ──────────────────────────────────────────
@router.get("/quick-bill", response_class=HTMLResponse)
def quick_bill_page(request: Request, q: str = None, db: Session = Depends(get_db)):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    store = ctx["store"]
    results = None
    if q:
        results = (
            db.query(Member)
            .filter(
                Member.store_id == store.id,
                (
                    Member.name.ilike(f"%{q}%")
                    | Member.phone.ilike(f"%{q}%")
                    | Member.member_code.ilike(f"%{q}%")
                    | Member.company_name.ilike(f"%{q}%")
                    | Member.license_plate.ilike(f"%{q}%")
                    | Member.driver_name.ilike(f"%{q}%")
                ),
            )
            .order_by(Member.id.desc())
            .limit(30)
            .all()
        )
    return templates.TemplateResponse(request, "quick_bill.html", {
        "store_name": store.name,
        "user_role": ctx["role"],
        "q": q,
        "results": results,
    })


@router.get("/cash-bill", response_class=HTMLResponse)
def cash_bill_page(request: Request, db: Session = Depends(get_db)):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    store = ctx["store"]
    products = _get_products(db, store.id)
    return templates.TemplateResponse(request, "cash_bill.html", {
        "store_name": store.name,
        "store": store,
        "user_role": ctx["role"],
        "products": products,
    })


@router.post("/cash-bill", response_class=HTMLResponse)
def cash_bill_post(
    request: Request,
    items_json: str = Form(...),
    customer_name: str = Form(""),
    payment_method: str = Form("cash"),
    note: str = Form(""),
    vat_type: str = Form("none"),
    db: Session = Depends(get_db),
):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    store = ctx["store"]

    items = json.loads(items_json)
    item_total = sum(float(i["qty"]) * float(i["price"]) for i in items)
    vat_rate = store.vat_rate or 7.0

    if vat_type == "exclusive":
        subtotal = round(item_total, 2)
        vat_amount = round(subtotal * vat_rate / 100, 2)
        total = round(subtotal + vat_amount, 2)
    elif vat_type == "inclusive":
        total = round(item_total, 2)
        subtotal = round(total / (1 + vat_rate / 100), 2)
        vat_amount = round(total - subtotal, 2)
    else:
        subtotal = round(item_total, 2)
        vat_amount = 0.0
        vat_rate = 0.0
        total = subtotal

    tx = Transaction(
        store_id=store.id,
        total=total,
        payment_method=payment_method,
        member_id=None,
        terminal_id="web-cash",
        raw={"items": items, "subtotal": subtotal, "vat_amount": vat_amount,
             "vat_rate": vat_rate, "vat_type": vat_type, "total": total,
             "payment_method": payment_method, "note": note,
             "customer_name": customer_name.strip() or None},
    )
    db.add(tx)
    db.commit()
    db.refresh(tx)

    receipt = Receipt(
        transaction_id=tx.id,
        raw_payload={
            "items": items,
            "subtotal": subtotal,
            "vat_amount": vat_amount,
            "vat_rate": vat_rate,
            "vat_type": vat_type,
            "total": total,
            "payment_method": payment_method,
            "note": note,
            "customer_name": customer_name.strip() or None,
            "points_earned": 0,
            "billing_profile": None,
            "is_cash": True,
        },
        created_by_name=ctx["actor_name"],
        created_by_id=ctx["actor_id"],
    )
    db.add(receipt)
    db.commit()
    db.refresh(receipt)

    return RedirectResponse(f"/web/receipts/{receipt.id}", status_code=302)



@router.get("/vat-bill", response_class=HTMLResponse)
def vat_bill_page(request: Request, db: Session = Depends(get_db)):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    store = ctx["store"]
    products = _get_products(db, store.id)

    # ใช้สมาชิกประเภท company เป็น dropdown หลัก
    company_members = (
        db.query(Member)
        .filter(Member.store_id == store.id, Member.member_type == "company")
        .order_by(Member.company_name, Member.name)
        .all()
    )
    # รวม billing_profiles จากสมาชิกทั่วไปด้วย (backward compat)
    seen_names = set()
    company_choices = []
    for m in company_members:
        display = m.company_name or m.name
        key = display.strip().lower()
        seen_names.add(key)
        company_choices.append({
            "value": f"mem_{m.id}",
            "company_name": display,
            "tax_id": m.tax_id,
            "address": m.address or "",
            "phone": m.phone or "",
            "email": m.email or "",
        })
    # billing_profiles ที่มีอยู่เดิม (ไม่ซ้ำชื่อ)
    member_ids = db.query(Member.id).filter(Member.store_id == store.id).scalar_subquery()
    bps = (
        db.query(BillingProfile)
        .filter(BillingProfile.member_id.in_(member_ids))
        .order_by(BillingProfile.company_name)
        .all()
    )
    for bp in bps:
        if not bp.company_name:
            continue
        key = bp.company_name.strip().lower()
        if key in seen_names:
            continue
        seen_names.add(key)
        company_choices.append({
            "value": f"bp_{bp.id}",
            "company_name": bp.company_name,
            "tax_id": bp.tax_id,
            "address": bp.address or "",
            "phone": bp.phone or "",
            "email": getattr(bp, "email", "") or "",
        })
    company_choices.sort(key=lambda x: (x["company_name"] or "").lower())

    return templates.TemplateResponse(request, "vat_bill.html", {
        "store_name": store.name,
        "store": store,
        "user_role": ctx["role"],
        "products": products,
        "company_choices": company_choices,
    })


@router.post("/vat-bill", response_class=HTMLResponse)
def vat_bill_post(
    request: Request,
    items_json: str = Form(...),
    payment_method: str = Form("cash"),
    note: str = Form(""),
    vat_type: str = Form("exclusive"),
    company_id: str = Form(""),
    driver_name: str = Form(""),
    license_plate: str = Form(""),
    new_company_name: str = Form(""),
    new_company_tax_id: str = Form(""),
    new_company_address: str = Form(""),
    new_company_phone: str = Form(""),
    new_company_email: str = Form(""),
    new_company_save: str = Form(""),
    override_address: str = Form(""),
    override_phone: str = Form(""),
    override_save: str = Form(""),
    db: Session = Depends(get_db),
):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    store = ctx["store"]

    items = json.loads(items_json)
    item_total = sum(float(i["qty"]) * float(i["price"]) for i in items)
    vat_rate = store.vat_rate or 7.0

    if vat_type == "exclusive":
        subtotal = round(item_total, 2)
        vat_amount = round(subtotal * vat_rate / 100, 2)
        total = round(subtotal + vat_amount, 2)
    elif vat_type == "inclusive":
        total = round(item_total, 2)
        subtotal = round(total / (1 + vat_rate / 100), 2)
        vat_amount = round(total - subtotal, 2)
    else:
        subtotal = round(item_total, 2)
        vat_amount = 0.0
        vat_rate = 0.0
        total = subtotal

    # Resolve billing profile snapshot from selected / new company
    bp_snapshot = None
    vat_member_id = None  # เชื่อม transaction กับสมาชิกประเภทบริษัท
    if company_id == "__new__" and new_company_name.strip():
        bp_snapshot = {
            "company_name": new_company_name.strip(),
            "tax_id": new_company_tax_id.strip() or None,
            "address": new_company_address.strip() or None,
            "phone": new_company_phone.strip() or None,
            "email": new_company_email.strip() or None,
            "label": new_company_name.strip(),
        }
        if new_company_save:
            # บันทึกเป็น Member type=company
            new_m = Member(
                member_type="company",
                name=new_company_name.strip(),
                company_name=new_company_name.strip(),
                tax_id=new_company_tax_id.strip() or None,
                address=new_company_address.strip() or None,
                phone=new_company_phone.strip() or None,
                email=new_company_email.strip() or None,
                store_id=store.id,
                tier="general",
                points=0,
            )
            db.add(new_m)
            db.flush()
            new_m.member_code = f"C{store.id:04d}{new_m.id:07d}"
            vat_member_id = new_m.id
    elif company_id.startswith("mem_"):
        try:
            m = db.query(Member).filter(
                Member.id == int(company_id[4:]),
                Member.store_id == store.id,
            ).first()
            if m:
                vat_member_id = m.id
                bp_snapshot = {
                    "company_name": m.company_name or m.name,
                    "tax_id": m.tax_id,
                    "address": m.address,
                    "phone": m.phone,
                    "email": m.email,
                    "label": m.company_name or m.name,
                }
        except (ValueError, Exception):
            pass
    elif company_id.startswith("co_"):
        try:
            co = db.query(Company).filter(
                Company.id == int(company_id[3:]),
                Company.store_id == store.id,
            ).first()
            if co:
                bp_snapshot = {
                    "company_name": co.company_name,
                    "tax_id": co.tax_id,
                    "address": co.address,
                    "phone": co.phone,
                    "email": co.email,
                    "label": co.company_name,
                }
        except (ValueError, Exception):
            pass
    elif company_id.startswith("bp_"):
        try:
            member_ids = db.query(Member.id).filter(Member.store_id == store.id).scalar_subquery()
            bp = db.query(BillingProfile).filter(
                BillingProfile.id == int(company_id[3:]),
                BillingProfile.member_id.in_(member_ids),
            ).first()
            if bp:
                bp_snapshot = {
                    "company_name": bp.company_name,
                    "tax_id": bp.tax_id,
                    "address": bp.address,
                    "phone": bp.phone,
                    "email": getattr(bp, "email", None),
                    "label": bp.company_name or bp.label,
                }
        except (ValueError, Exception):
            pass

    # Apply override address/phone for existing companies
    if bp_snapshot is not None and company_id != "__new__":
        if override_address.strip():
            bp_snapshot["address"] = override_address.strip()
        if override_phone.strip():
            bp_snapshot["phone"] = override_phone.strip()
        # Save overridden address back to Member record if requested
        if override_save and company_id.startswith("mem_"):
            try:
                m_upd = db.query(Member).filter(
                    Member.id == int(company_id[4:]),
                    Member.store_id == store.id,
                ).first()
                if m_upd:
                    if override_address.strip():
                        m_upd.address = override_address.strip()
                    if override_phone.strip():
                        m_upd.phone = override_phone.strip()
                    db.flush()
            except (ValueError, Exception):
                pass

    clean_plate = license_plate.strip().upper() if license_plate and license_plate.strip() else None
    clean_driver = driver_name.strip() if driver_name and driver_name.strip() else None

    tx = Transaction(
        store_id=store.id,
        total=total,
        payment_method=payment_method,
        member_id=vat_member_id,
        terminal_id="web-vat",
        raw={"items": items, "subtotal": subtotal, "vat_amount": vat_amount,
             "vat_rate": vat_rate, "vat_type": vat_type, "total": total,
             "payment_method": payment_method, "note": note},
    )
    db.add(tx)
    db.commit()
    db.refresh(tx)

    receipt = Receipt(
        transaction_id=tx.id,
        raw_payload={
            "bill_type": "vat",
            "items": items,
            "subtotal": subtotal,
            "vat_amount": vat_amount,
            "vat_rate": vat_rate,
            "vat_type": vat_type,
            "total": total,
            "payment_method": payment_method,
            "note": note,
            "points_earned": 0,
            "driver_name": clean_driver,
            "license_plate": clean_plate,
            "billing_profile": bp_snapshot,
            "is_cash": False,
        },
        created_by_name=ctx["actor_name"],
        created_by_id=ctx["actor_id"],
    )
    db.add(receipt)
    db.commit()
    db.refresh(receipt)

    return RedirectResponse(f"/web/receipts/{receipt.id}", status_code=302)



@router.get("/dashboard", response_class=HTMLResponse)
def dashboard_page(request: Request, db: Session = Depends(get_db)):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    store = ctx["store"]

    now = datetime.now()
    month_start = datetime(now.year, now.month, 1)

    total_members = db.query(Member).filter(Member.store_id == store.id).count()
    new_members_month = db.query(Member).filter(
        Member.store_id == store.id,
        Member.created_at >= month_start,
    ).count()

    total_tx = db.query(Transaction).filter(
        Transaction.store_id == store.id
    ).count()
    tx_month = db.query(Transaction).filter(
        Transaction.store_id == store.id,
        Transaction.timestamp >= month_start,
    ).count()

    revenue_month = db.query(func.coalesce(func.sum(Transaction.total), 0)).filter(
        Transaction.store_id == store.id,
        Transaction.timestamp >= month_start,
    ).scalar()

    staff_count = db.query(StaffUser).filter(StaffUser.store_id == store.id, StaffUser.is_active == True).count()
    product_count = db.query(Product).filter(Product.store_id == store.id, Product.is_active == True).count()

    limits = _get_plan_limits(store, db)

    # Recent transactions
    recent_tx = (
        db.query(Transaction)
        .filter(Transaction.store_id == store.id)
        .order_by(Transaction.timestamp.desc())
        .limit(10)
        .all()
    )

    # Top 5 members by points
    top_members = (
        db.query(Member)
        .filter(Member.store_id == store.id)
        .order_by(Member.points.desc())
        .limit(5)
        .all()
    )

    return templates.TemplateResponse(request, "dashboard.html", {
        "store_name": store.name,
        "user_role": ctx["role"],
        "total_members": total_members,
        "new_members_month": new_members_month,
        "total_tx": total_tx,
        "tx_month": tx_month,
        "revenue_month": revenue_month,
        "staff_count": staff_count,
        "product_count": product_count,
        "limits": limits,
        "recent_tx": recent_tx,
        "top_members": top_members,
    })


# ──────────────────────────────────────────
# Password change
# ──────────────────────────────────────────
@router.get("/change-password", response_class=HTMLResponse)
def change_password_page(request: Request, db: Session = Depends(get_db)):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)
    return templates.TemplateResponse(request, "change_password.html", {
        "store_name": ctx["store"].name,
        "user_role": ctx["role"],
    })


@router.post("/change-password")
def change_password_post(
    request: Request,
    current_password: str = Form(...),
    new_password: str = Form(...),
    confirm_password: str = Form(...),
    db: Session = Depends(get_db),
):
    ctx = _get_user_ctx(request, db)
    if not ctx:
        return RedirectResponse("/web/login", status_code=302)

    def _err(msg):
        return templates.TemplateResponse(request, "change_password.html", {
            "store_name": ctx["store"].name,
            "user_role": ctx["role"],
            "error": msg,
        })

    if new_password != confirm_password:
        return _err("รหัสผ่านใหม่ไม่ตรงกัน")

    pwd_error = validate_password(new_password)
    if pwd_error:
        return _err(pwd_error)

    # Determine which entity to update
    if ctx["role"] == "owner":
        store = ctx["store"]
        if not verify_password(current_password, store.hashed_password):
            return _err("รหัสผ่านปัจจุบันไม่ถูกต้อง")
        store.hashed_password = hash_password(new_password)
        db.add(store)
    else:
        staff = db.query(StaffUser).filter(StaffUser.id == ctx["actor_id"]).first()
        if not staff or not verify_password(current_password, staff.hashed_password):
            return _err("รหัสผ่านปัจจุบันไม่ถูกต้อง")
        staff.hashed_password = hash_password(new_password)
        db.add(staff)

    db.commit()
    return templates.TemplateResponse(request, "change_password.html", {
        "store_name": ctx["store"].name,
        "user_role": ctx["role"],
        "success": "เปลี่ยนรหัสผ่านสำเร็จ",
    })
